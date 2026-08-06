# -*- coding: utf-8 -*-
"""data.json → 多 sheet 单工作簿 Excel：SKU明细 / 商品汇总 / 核对总览 / 差异明细。
用法: python3 build_report.py <data.json> <out.xlsx> [--manifest imgs/manifest.json] [--shop 店铺名]
--manifest 提供时内嵌 SKU 图片（url→本地缩略图映射，相对路径按 manifest 所在目录解析）。"""
import argparse, json, os, re, datetime, difflib
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

HEADER_FILL = PatternFill("solid", fgColor="0284C7")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
NOTE_FONT = Font(name="Arial", size=9, italic=True, color="808080")
WARN_FONT = Font(name="Arial", size=9, color="9C6500")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE"); BAD_FONT = Font(name="Arial", size=10, bold=True, color="9C0006")
OK_FILL = PatternFill("solid", fgColor="C6EFCE"); OK_FONT = Font(name="Arial", size=10, color="006100")
SOLO_FILL = PatternFill("solid", fgColor="F2F2F2"); SOLO_FONT = Font(name="Arial", size=10, color="595959")
DEV_FILL = PatternFill("solid", fgColor="FFC7CE"); DEV_FONT = Font(name="Arial", size=10, bold=True, color="9C0006")
THIN = Side(style="thin", color="D9E2EC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
IMG_PX, ROW_PT = 62, 50

def norm(s):
    return "".join(ch for ch in s if ch.isalnum())

def mode_price(rows, key):
    c = Counter(r[key] for r in rows if r.get(key) is not None)
    if not c:
        return None, c
    top = c.most_common()
    best = sorted([k for k, v in top if v == top[0][1]], key=lambda x: float(x))
    return best[0], c

def fmt_range(counter):
    ks = sorted(counter.keys(), key=lambda x: float(x))
    return ks[0] if len(ks) == 1 else " / ".join(f"{k}×{counter[k]}" for k in ks)

def build_groups(data):
    """同一 SKU 图片 URL = 同一真实 SKU（卖家多链接复用同款 SKU 图）。"""
    groups = {}
    for item in data.get("items", []):
        for row in item.get("rows", []):
            key = row.get("image") or f"_noimg_{item['itemId']}_{row['skuId']}"
            groups.setdefault(key, []).append(
                {"itemId": item["itemId"], "title": item["title"], "url": item["url"], **row})
    sku_list = []
    for img, rows in groups.items():
        promo_mode, promo_c = mode_price(rows, "promoPrice")
        list_mode, list_c = mode_price(rows, "listPrice")
        name_mode = Counter(r["skuName"] for r in rows).most_common(1)[0][0]
        sku_list.append(dict(image=img if not img.startswith("_noimg_") else "", name=name_mode,
                             rows=rows, n=len(rows), promo_c=promo_c, list_c=list_c,
                             promo_mode=promo_mode, list_mode=list_mode,
                             promo_conflict=len(promo_c) > 1, list_conflict=len(list_c) > 1,
                             conflict=len(promo_c) > 1 or len(list_c) > 1))
    sku_list.sort(key=lambda s: (-s["conflict"], -s["n"]))
    return sku_list

def build_warnings(sku_list):
    """疑似同 SKU：同容量+高相似但图片不同；排除配置(wifi/旋钮/触控)与颜色差异。"""
    warnings = []
    for a in sku_list:
        for b in sku_list:
            if a["image"] >= b["image"] or not a["image"] or not b["image"]:
                continue
            na, nb = norm(a["name"]), norm(b["name"])
            capa = re.findall(r"\d+L", na); capb = re.findall(r"\d+L", nb)
            if not capa or capa != capb:
                continue
            if ("wifi" in na.lower()) != ("wifi" in nb.lower()):
                continue
            if ("旋钮" in na) != ("旋钮" in nb) or ("触控" in na) != ("触控" in nb):
                continue
            colors = ["枪灰色", "白色", "灰色", "金色", "黑色"]
            ca = {c for c in colors if c in na}; cb = {c for c in colors if c in nb}
            if ca and cb and ca != cb:
                continue
            if difflib.SequenceMatcher(None, na, nb).ratio() < 0.8:
                continue
            pa = a["promo_mode"] if a["n"] > 1 else a["rows"][0]["promoPrice"]
            pb = b["promo_mode"] if b["n"] > 1 else b["rows"][0]["promoPrice"]
            same = "价格一致" if pa == pb else f"价格不一致({pa} vs {pb})"
            warnings.append(f"「{a['name']}」(实付{pa}) 与 「{b['name']}」(实付{pb}) "
                            f"名称高度相似但 SKU 图不同，疑似同一真实 SKU，{same}，请人工确认。")
    return warnings

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, CENTER, BORDER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("data"); ap.add_argument("out")
    ap.add_argument("--manifest", default=None); ap.add_argument("--shop", default=None)
    args = ap.parse_args()

    with open(args.data, encoding="utf-8") as f:
        data = json.load(f)
    manifest, mdir = {}, os.path.dirname(os.path.abspath(args.manifest)) if args.manifest else ""
    if args.manifest and os.path.exists(args.manifest):
        with open(args.manifest, encoding="utf-8") as f:
            manifest = json.load(f)

    def embed(ws, url, ref):
        p = manifest.get(url)
        if not p:
            return
        cand = [p] if os.path.isabs(p) else [os.path.join(mdir, p), os.path.join(os.getcwd(), p)]
        for c in cand:
            if os.path.exists(c):
                im = XLImage(c); im.width, im.height = IMG_PX, IMG_PX
                ws.add_image(im, ref)
                return

    shop = args.shop or data.get("shop", "")
    sku_list = build_groups(data)
    warnings = build_warnings(sku_list)
    wb = Workbook()

    # ---------- Sheet1 SKU明细 ----------
    ws = wb.active; ws.title = "SKU明细"
    ws.append(["商品ID", "SKU图片", "SKU ID", "商品标题", "SKU名称", "原价(元)", "实付价(元)", "库存", "图片链接"])
    style_header(ws, 9)
    r = 2
    for item in data.get("items", []):
        for row in item.get("rows", []):
            ws.cell(row=r, column=1, value=item["itemId"])
            ws.cell(row=r, column=3, value=row["skuId"])
            ws.cell(row=r, column=4, value=item["title"]).hyperlink = item["url"]
            ws.cell(row=r, column=5, value=row["skuName"])
            ws.cell(row=r, column=6, value=float(row["listPrice"])).number_format = "#,##0.00"
            ws.cell(row=r, column=7, value=float(row["promoPrice"])).number_format = "#,##0.00"
            ws.cell(row=r, column=8, value=int(row["quantity"]) if row.get("quantity") is not None else None)
            ws.cell(row=r, column=9, value=row["image"]).hyperlink = row["image"] or None
            ws.row_dimensions[r].height = ROW_PT
            embed(ws, row["image"], f"B{r}")
            r += 1
    for cells in ws.iter_rows(min_row=2, max_row=r - 1, min_col=1, max_col=9):
        for cell in cells:
            cell.border = BORDER
            cell.font = LINK_FONT if cell.column in (4, 9) else BODY_FONT
            cell.alignment = CENTER if cell.column in (1, 3, 6, 7, 8) else LEFT
    for i, w in enumerate([16, 10, 16, 44, 40, 10, 10, 8, 56], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:I{r-1}"

    # ---------- Sheet2 商品汇总 ----------
    ws2 = wb.create_sheet("商品汇总")
    ws2.append(["商品ID", "商品标题", "主图", "SKU数量", "实付最低价(元)", "实付最高价(元)", "原价区间(元)"])
    style_header(ws2, 7)
    r = 2
    for item in data.get("items", []):
        promos = [float(x["promoPrice"]) for x in item["rows"] if x.get("promoPrice")]
        lists = [float(x["listPrice"]) for x in item["rows"] if x.get("listPrice")]
        ws2.cell(row=r, column=1, value=item["itemId"])
        ws2.cell(row=r, column=2, value=item["title"]).hyperlink = item["url"]
        ws2.cell(row=r, column=4, value=len(item["rows"]))
        ws2.cell(row=r, column=5, value=min(promos)).number_format = "#,##0.00"
        ws2.cell(row=r, column=6, value=max(promos)).number_format = "#,##0.00"
        ws2.cell(row=r, column=7, value=f"{min(lists):.0f} - {max(lists):.0f}")
        ws2.row_dimensions[r].height = ROW_PT
        embed(ws2, item["rows"][0]["image"], f"C{r}")
        r += 1
    for cells in ws2.iter_rows(min_row=2, max_row=r - 1, min_col=1, max_col=7):
        for cell in cells:
            cell.border = BORDER
            cell.font = LINK_FONT if cell.column == 2 else BODY_FONT
            cell.alignment = CENTER if cell.column in (1, 4, 5, 6, 7) else LEFT
    for i, w in enumerate([16, 44, 10, 10, 14, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.auto_filter.ref = f"A1:G{r-1}"

    # ---------- Sheet3 核对总览 ----------
    ws3 = wb.create_sheet("核对总览")
    ws3.append(["SKU图片", "真实SKU名称(代表名)", "出现链接数", "一致性状态", "实付价分布", "原价分布", "建议实付价(众数)", "差异说明"])
    style_header(ws3, 8)
    r = 2
    for s in sku_list:
        ws3.cell(row=r, column=2, value=s["name"])
        ws3.cell(row=r, column=3, value=s["n"])
        status = ws3.cell(row=r, column=4)
        if s["conflict"]:
            status.value = "⚠ 价格不一致"; status.fill, status.font = BAD_FILL, BAD_FONT
            parts = []
            if s["promo_conflict"]:
                devs = [f"{x['itemId']}实付{x['promoPrice']}" for x in s["rows"] if x["promoPrice"] != s["promo_mode"]]
                parts.append("实付价偏离：" + "；".join(devs))
            if s["list_conflict"]:
                devs = [f"{x['itemId']}原价{x['listPrice']}" for x in s["rows"] if x["listPrice"] != s["list_mode"]]
                parts.append("原价偏离：" + "；".join(devs))
            ws3.cell(row=r, column=8, value="（基准=众数价）" + " ｜ ".join(parts))
        elif s["n"] == 1:
            status.value = "仅1个链接"; status.fill, status.font = SOLO_FILL, SOLO_FONT
        else:
            status.value = "✓ 一致"; status.fill, status.font = OK_FILL, OK_FONT
        ws3.cell(row=r, column=5, value=fmt_range(s["promo_c"]))
        ws3.cell(row=r, column=6, value=fmt_range(s["list_c"]))
        if s["promo_mode"]:
            ws3.cell(row=r, column=7, value=float(s["promo_mode"])).number_format = "#,##0.00"
        ws3.row_dimensions[r].height = ROW_PT
        embed(ws3, s["image"], f"A{r}")
        r += 1
    for cells in ws3.iter_rows(min_row=2, max_row=r - 1, min_col=1, max_col=8):
        for cell in cells:
            cell.border = BORDER
            if cell.column != 4:
                cell.font = BODY_FONT
            cell.alignment = LEFT if cell.column in (2, 5, 6, 8) else CENTER
    for i, w in enumerate([10, 36, 10, 13, 18, 18, 12, 62], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    nr = r + 1
    shop_txt = f"（店铺：{shop}）" if shop else ""
    ws3.cell(row=nr, column=1, value=f"采集时间：{datetime.date.today().isoformat()}（Asia/Shanghai）{shop_txt}｜判定口径：SKU 图片相同 = 同一真实 SKU；基准价取各链接中的众数价，偏离众数的价格标红。价格为采集时点页面展示价（原价=优惠前，实付价=店铺优惠后）。").font = NOTE_FONT
    ws3.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=8)
    for i, wtext in enumerate(warnings, 1):
        ws3.cell(row=nr + i, column=1, value=f"疑似同SKU提示{i}：{wtext}").font = WARN_FONT
        ws3.merge_cells(start_row=nr + i, start_column=1, end_row=nr + i, end_column=8)

    # ---------- Sheet4 差异明细 ----------
    ws4 = wb.create_sheet("差异明细")
    ws4.append(["SKU图片", "真实SKU名称(代表名)", "商品ID", "商品标题", "该链接SKU ID", "该链接SKU名称", "原价(元)", "实付价(元)", "与建议价差额"])
    style_header(ws4, 9)
    r = 2; dev_cells = set()
    for s in [x for x in sku_list if x["conflict"]]:
        for row in sorted(s["rows"], key=lambda x: float(x["promoPrice"])):
            ws4.cell(row=r, column=2, value=s["name"])
            ws4.cell(row=r, column=3, value=row["itemId"])
            ws4.cell(row=r, column=4, value=row["title"]).hyperlink = row["url"]
            ws4.cell(row=r, column=5, value=row["skuId"])
            ws4.cell(row=r, column=6, value=row["skuName"])
            lp = ws4.cell(row=r, column=7, value=float(row["listPrice"])); lp.number_format = "#,##0.00"
            pp = ws4.cell(row=r, column=8, value=float(row["promoPrice"])); pp.number_format = "#,##0.00"
            diff = float(row["promoPrice"]) - float(s["promo_mode"])
            dc = ws4.cell(row=r, column=9, value=diff); dc.number_format = "+#,##0.00;-#,##0.00;0"
            if row["promoPrice"] != s["promo_mode"]:
                pp.fill, pp.font = DEV_FILL, DEV_FONT
                dc.fill, dc.font = DEV_FILL, DEV_FONT
                dev_cells |= {(r, 8), (r, 9)}
            if row["listPrice"] != s["list_mode"]:
                lp.fill, lp.font = DEV_FILL, DEV_FONT
                dev_cells.add((r, 7))
            ws4.row_dimensions[r].height = ROW_PT
            embed(ws4, s["image"], f"A{r}")
            r += 1
    for cells in ws4.iter_rows(min_row=2, max_row=r - 1, min_col=1, max_col=9):
        for cell in cells:
            cell.border = BORDER
            if (cell.row, cell.column) not in dev_cells:
                cell.font = LINK_FONT if cell.column == 4 else BODY_FONT
            cell.alignment = LEFT if cell.column in (2, 4, 6) else CENTER
    for i, w in enumerate([10, 30, 15, 40, 16, 34, 10, 10, 12], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.auto_filter.ref = f"A1:I{r-1}"
    ws4.cell(row=r + 1, column=1, value="标红单元格 = 偏离众数基准价的链接。同一真实 SKU 按价格升序排列，点击商品标题可直达对应链接改价。").font = NOTE_FONT
    ws4.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=9)

    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    wb.save(args.out)
    total = sum(len(i["rows"]) for i in data.get("items", []))
    print(json.dumps({"out": args.out, "items": len(data.get("items", [])), "skus": total,
                      "real_skus": len(sku_list), "conflict": sum(1 for s in sku_list if s["conflict"]),
                      "warnings": len(warnings)}, ensure_ascii=False))

if __name__ == "__main__":
    main()
