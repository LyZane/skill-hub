# -*- coding: utf-8 -*-
"""跨平台比价：天猫真实 SKU（按图片归组、众数实付价）vs 京东链接（每页一个 SKU）。
用法: python3 build_compare.py <tmall_data.json> <jd_data.json> <out.xlsx> [--manifest imgs/manifest.json]
匹配规则：关键词特征重合度 + 标题相似度打分；容量冲突惩罚；低置信度标黄待人工复核。"""
import argparse, json, os, re, datetime, difflib
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

HEADER_FILL = PatternFill("solid", fgColor="0284C7")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
NOTE_FONT = Font(name="Arial", size=9, italic=True, color="808080")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C"); WARN_FONT = Font(name="Arial", size=10, color="9C6500")
HI_FILL = PatternFill("solid", fgColor="FFC7CE"); HI_FONT = Font(name="Arial", size=10, bold=True, color="9C0006")
LO_FILL = PatternFill("solid", fgColor="C6EFCE"); LO_FONT = Font(name="Arial", size=10, color="006100")
THIN = Side(style="thin", color="D9E2EC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
IMG_PX, ROW_PT = 62, 50

KEYWORDS = {
    "无雾": ["无雾", "冷蒸发", "蒸发式"],
    "WiFi": ["wifi"],
    "触控": ["触控"],
    "机械": ["机械", "旋钮"],
    "语音": ["语音", "声控"],
    "冷热雾": ["冷热雾", "冷暖", "热雾", "冷雾"],
    "负离子": ["负离子"],
    "除菌": ["除菌", "杀菌", "uv"],
    "台式": ["台面", "桌面", "台式"],
    "落地": ["落地", "柜式"],
    "香薰": ["香薰"],
    "夜灯": ["夜灯", "氛围灯", "七彩灯", "小夜灯"],
    "恒湿": ["恒湿"],
    "净化": ["净化"],
}

def norm(s):
    return "".join(ch for ch in (s or "").lower() if ch.isalnum())

def feats(name):
    n = norm(name)
    return {k for k, alts in KEYWORDS.items() if any(norm(a) in n for a in alts)}

def capacity(name):
    m = re.findall(r"(\d+)\s*l(?![a-z])", (name or "").lower())
    return m[0] if m else None

def score(tmall_text, jd_text, tm_price, jd_price):
    ft, fj = feats(tmall_text), feats(jd_text)
    jac = (2 * len(ft & fj) / (len(ft) + len(fj))) if (ft or fj) else 0.0
    ratio = difflib.SequenceMatcher(None, norm(tmall_text), norm(jd_text)).ratio()
    price_prox = 0.0
    if tm_price and jd_price:
        hi = max(float(tm_price), float(jd_price))
        price_prox = 1 - abs(float(tm_price) - float(jd_price)) / hi if hi else 0
    s = 0.35 * jac + 0.20 * ratio + 0.45 * price_prox
    ct, cj = capacity(tmall_text), capacity(jd_text)
    if ct and cj and ct != cj:
        s *= 0.3
    return min(s, 1.0)

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, CENTER, BORDER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("tmall"); ap.add_argument("jd"); ap.add_argument("out")
    ap.add_argument("--manifest", default=None)
    args = ap.parse_args()

    with open(args.tmall, encoding="utf-8") as f:
        tm = json.load(f)
    with open(args.jd, encoding="utf-8") as f:
        jd = json.load(f)
    manifest, mdir = {}, os.path.dirname(os.path.abspath(args.manifest)) if args.manifest else ""
    if args.manifest and os.path.exists(args.manifest):
        with open(args.manifest, encoding="utf-8") as f:
            manifest = json.load(f)

    def embed(ws, url, ref):
        p = manifest.get(url)
        if not p:
            return
        cand = [p] if os.path.isabs(p) else [os.path.join(mdir, p), os.path.join(os.getcwd(), p)]
        for c in cand:
            if os.path.exists(c):
                im = XLImage(c); im.width, im.height = IMG_PX, IMG_PX
                ws.add_image(im, ref)
                return

    # 天猫真实 SKU 归组（图片 URL 相同）
    groups = []
    seen = {}
    for item in tm.get("items", []):
        for row in item.get("rows", []):
            key = row.get("image") or f"_noimg_{item['itemId']}_{row['skuId']}"
            g = seen.get(key)
            if not g:
                g = dict(image=key if not key.startswith("_noimg_") else "", rows=[], names=[],
                         title=item.get("title", ""))
                seen[key] = g; groups.append(g)
            g["rows"].append({**row, "itemId": item["itemId"], "url": item["url"]})
            g["names"].append(row["skuName"])
    for g in groups:
        g["name"] = Counter(g["names"]).most_common(1)[0][0]
        g["text"] = g["name"] + " " + g["title"]
        c = Counter(r["promoPrice"] for r in g["rows"] if r.get("promoPrice"))
        top = c.most_common()
        g["price"] = sorted([k for k, v in top if v == top[0][1]], key=lambda x: float(x))[0] if top else None
        g["n"] = len(g["rows"])

    # 京东 SKU（每链接一行）
    jd_rows = []
    for item in jd.get("items", []):
        for row in item.get("rows", []):
            jd_rows.append({**row, "jdTitle": item["title"], "url": item["url"],
                            "commentCount": item.get("commentCount"), "shop": item.get("shop")})

    # 匹配：关键词重合 + 名称相似 + 价格接近
    matched, review, used = [], [], set()
    for jr in jd_rows:
        jd_name = jr["skuName"] or jr["jdTitle"]
        jd_text = jd_name + " " + jr["jdTitle"]
        cand = sorted(((score(g["text"], jd_text, g["price"], jr["promoPrice"]), i)
                       for i, g in enumerate(groups)), reverse=True)
        s, i = cand[0]
        s2 = cand[1][0] if len(cand) > 1 else 0
        conf = "高" if (s >= 0.8 and s - s2 >= 0.06) else ("中" if s >= 0.55 else "低")
        rec = dict(jr=jr, g=groups[i], score=s, conf=conf,
                   alt=[(round(x, 2), groups[gi]["name"]) for x, gi in cand[1:3]])
        if conf == "高":
            if i in used:
                rec["conf"] = "中"; rec["note"] = "该天猫SKU已被其他京东链接高置信匹配"
                review.append(rec)
            else:
                used.add(i); matched.append(rec)
        else:
            if i in used:
                rec["note"] = "该天猫SKU已被其他京东链接高置信匹配"
            review.append(rec)
    unmatched_tm = [g for i, g in enumerate(groups) if i not in used]

    wb = Workbook()
    ws = wb.active; ws.title = "比价总览"
    ws.append(["天猫SKU图", "天猫真实SKU(代表名)", "天猫实付价(众数)", "天猫链接数",
               "京东规格名", "京东实付价", "京东划线价", "京东评价数", "差额(天-京)", "谁更低", "置信度", "京东链接"])
    style_header(ws, 12)
    r = 2
    for rec in sorted(matched, key=lambda x: x["g"]["name"]):
        g, jr = rec["g"], rec["jr"]
        ws.cell(row=r, column=2, value=g["name"])
        ws.cell(row=r, column=3, value=float(g["price"])).number_format = "#,##0.00"
        ws.cell(row=r, column=4, value=g["n"])
        ws.cell(row=r, column=5, value=jr["skuName"] or jr["jdTitle"])
        ws.cell(row=r, column=6, value=float(jr["promoPrice"])).number_format = "#,##0.00"
        if jr.get("listPrice"):
            ws.cell(row=r, column=7, value=float(jr["listPrice"])).number_format = "#,##0.00"
        ws.cell(row=r, column=8, value=jr.get("commentCount"))
        diff = float(g["price"]) - float(jr["promoPrice"])
        dc = ws.cell(row=r, column=9, value=diff); dc.number_format = "+#,##0.00;-#,##0.00;0"
        cmp_cell = ws.cell(row=r, column=10, value="天猫更高" if diff > 0 else ("京东更高" if diff < 0 else "持平"))
        if diff > 0: cmp_cell.fill, cmp_cell.font = HI_FILL, HI_FONT
        elif diff < 0: cmp_cell.fill, cmp_cell.font = LO_FILL, LO_FONT
        ws.cell(row=r, column=11, value=f"{rec['conf']}({rec['score']:.2f})")
        ws.cell(row=r, column=12, value=jr["url"]).hyperlink = jr["url"]
        ws.row_dimensions[r].height = ROW_PT
        embed(ws, g["image"], f"A{r}")
        r += 1
    for cells in ws.iter_rows(min_row=2, max_row=r - 1, min_col=1, max_col=12):
        for cell in cells:
            cell.border = BORDER
            if not cell.font or cell.font.color is None or cell.font.color.rgb in (None, "FF000000"):
                cell.font = LINK_FONT if cell.column == 12 else BODY_FONT
            cell.alignment = LEFT if cell.column in (2, 5, 12) else CENTER
    for i, w in enumerate([10, 34, 12, 9, 30, 10, 10, 9, 11, 9, 10, 42], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:L{r-1}"
    last = r
    note = (f"采集时间：{datetime.date.today().isoformat()}（Asia/Shanghai）｜天猫价=同店各链接实付价众数；京东价=页面到手价（受登录账号/收货地区影响）。"
            "差额=天猫-京东，正值（红色）=天猫卖得更贵，负值（绿色）=京东更贵。")
    ws.cell(row=last + 1, column=1, value=note).font = NOTE_FONT
    ws.merge_cells(start_row=last + 1, start_column=1, end_row=last + 1, end_column=12)

    ws2 = wb.create_sheet("待复核与未匹配")
    ws2.append(["类型", "名称", "价格", "最可能对应(天猫/京东)", "相似度", "备注"])
    style_header(ws2, 6)
    r = 2
    for rec in sorted(review, key=lambda x: -x["score"]):
        jr, g = rec["jr"], rec["g"]
        ws2.cell(row=r, column=1, value="京东链接(待复核)")
        ws2.cell(row=r, column=2, value=(jr["skuName"] or jr["jdTitle"]) + " ｜ " + jr["jdTitle"][:30])
        ws2.cell(row=r, column=3, value=float(jr["promoPrice"])).number_format = "#,##0.00"
        ws2.cell(row=r, column=4, value=g["name"] + f"（天猫实付{g['price']}）")
        ws2.cell(row=r, column=5, value=round(rec["score"], 2))
        note = rec.get("note") or "置信度不足，请人工确认是否同款"
        if rec.get("alt"):
            note += "；备选：" + "；".join(f"{n}({s})" for s, n in rec["alt"])
        ws2.cell(row=r, column=6, value=note)
        for c in range(1, 7): ws2.cell(row=r, column=c).fill = WARN_FILL
        r += 1
    for g in unmatched_tm:
        ws2.cell(row=r, column=1, value="天猫SKU(无高置信对应)")
        ws2.cell(row=r, column=2, value=g["name"])
        ws2.cell(row=r, column=3, value=float(g["price"]) if g["price"] else None)
        ws2.cell(row=r, column=6, value=f"出现在 {g['n']} 个天猫链接")
        r += 1
    for cells in ws2.iter_rows(min_row=2, max_row=r - 1, min_col=1, max_col=6):
        for cell in cells:
            cell.border = BORDER
            if not cell.fill or cell.fill.fgColor.rgb in (None, "00000000"):
                pass
            cell.font = BODY_FONT
            cell.alignment = LEFT if cell.column in (2, 4, 6) else CENTER
    for i, w in enumerate([18, 52, 10, 40, 8, 30], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    wb.save(args.out)
    print(json.dumps({"out": args.out, "matched": len(matched), "review": len(review),
                      "unmatched_tmall": len(unmatched_tm), "jd_total": len(jd_rows),
                      "tmall_groups": len(groups)}, ensure_ascii=False))

if __name__ == "__main__":
    main()
