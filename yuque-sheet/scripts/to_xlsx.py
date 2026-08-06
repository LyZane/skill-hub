#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""data.json → xlsx。
用法: python3 to_xlsx.py <data.json> <out.xlsx> [--with-images <imgs_dir>]
写值（纯数字串转数值）、合并单元格、系列色带填充、表头加粗、列宽行高；
--with-images 时在第 2 行嵌入产品图（basename 取 img url 末段）。
依赖 openpyxl（本机未预装，缺则 pip3 install openpyxl）。
"""
import json, os, re, sys

def rgba_hex(s):
    m = re.findall(r"\d+", s or "")
    return "%02X%02X%02X" % tuple(int(x) for x in m[:3]) if len(m) >= 3 else None

def main():
    data_path, out_path = sys.argv[1], sys.argv[2]
    imgs_dir = None
    if "--with-images" in sys.argv:
        imgs_dir = sys.argv[sys.argv.index("--with-images") + 1]
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("缺少 openpyxl，请先执行: pip3 install openpyxl")

    data = json.load(open(data_path, encoding="utf-8"))
    cells, merges = data["cells"], data["merges"]
    back_colors = [rgba_hex(c) for c in data.get("backColors") or []]
    title = (data.get("title") or "Sheet")[:31]

    max_r = max_c = 0
    grid = {}
    for r, c, v, bg in cells:
        grid[(r, c)] = (v, bg)
        max_r, max_c = max(max_r, r), max(max_c, c)

    wb = Workbook()
    ws = wb.active
    ws.title = title
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def to_value(v):
        if isinstance(v, dict):
            return v.get("link") or v.get("img") or None
        if isinstance(v, str) and re.fullmatch(r"\d+(\.\d+)?", v):
            return float(v) if "." in v else int(v)
        return v

    for r in range(max_r + 1):
        for c in range(max_c + 1):
            cell = ws.cell(row=r + 1, column=c + 1)
            e = grid.get((r, c))
            bg = e[1] if e else -1
            if e is not None:
                v, bg = e
                if isinstance(v, dict) and "img" in v and imgs_dir:
                    fn = v["img"].split("/")[-1]
                    p = os.path.join(imgs_dir, fn)
                    if os.path.exists(p):
                        from openpyxl.drawing.image import Image as XlImage
                        try:
                            from PIL import Image as PILImage
                            with PILImage.open(p) as im:
                                w0, h0 = im.size
                            h = 84
                            img = XlImage(p)
                            img.width, img.height = w0 * h / h0, h
                        except Exception:
                            img = XlImage(p)
                            img.width, img.height = 84, 84
                        ws.add_image(img, f"{get_column_letter(c + 1)}{r + 1}")
                else:
                    cell.value = to_value(v)
            if bg >= 0 and bg < len(back_colors) and back_colors[bg]:
                cell.fill = PatternFill("solid", fgColor=back_colors[bg])
            cell.alignment = left if c == 0 else center
            if r == 0:
                cell.font = Font(bold=True)
            if c == 0 and r > 0:
                cell.font = Font(bold=True)

    for r, c, rs, cs in merges:
        ws.merge_cells(start_row=r + 1, start_column=c + 1, end_row=r + rs, end_column=c + cs)

    ws.column_dimensions["A"].width = 20
    for c in range(1, max_c + 1):
        ws.column_dimensions[get_column_letter(c + 1)].width = 14
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 96 if imgs_dir else 24

    wb.save(out_path)
    print("saved", out_path, f"({max_r + 1}x{max_c + 1}, merges={len(merges)})")

if __name__ == "__main__":
    main()
